from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.graphics.charts.barcharts import HorizontalBarChart
from reportlab.graphics.shapes import Circle, Drawing, Line, Rect, String
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from utils.version import APP_VERSION

BASE = Path(__file__).resolve().parents[1]
OUTPUT = BASE / "output"
ASSETS = BASE / "assets"
OUTPUT.mkdir(exist_ok=True)
ASSETS.mkdir(exist_ok=True)

DARK = "1F2933"
TEAL = "00A7A7"
CYAN = "19C7D1"
LIGHT = "F3F6F8"
MID = "D7E1E7"
TEXT = "263238"
WHITE = "FFFFFF"
GREEN = "2E7D32"
YELLOW = "F9A825"
ORANGE = "EF6C00"
RED = "C62828"
BLUE = "1976D2"
PURPLE = "7B1FA2"

STATUS_COLORS = {
    "NORMAL": GREEN,
    "ATENCAO": YELLOW,
    "CRITICO": ORANGE,
    "EMERGENCIA": RED,
}

RISK_COLORS = {
    "CONTROLADO": GREEN,
    "RISCO MODERADO": YELLOW,
    "RISCO ALTO": ORANGE,
    "ESTOURO PREVISTO": RED,
    "ESTOURADO": RED,
    "SEM FRANQUIA": "78909C",
    "DESABILITADO": "78909C",
}

CHART_COLORS = [TEAL, BLUE, ORANGE, PURPLE, GREEN, RED, "455A64", "8D6E63", "00838F", "5D4037"]


def _fmt_gb(v):
    if v is None:
        return "-"
    v = float(v)
    if v >= 1000:
        return f"{v/1000:.2f} TB"
    return f"{v:.1f} GB"


def _fmt_date(value):
    if not value:
        return "-"
    try:
        return datetime.fromisoformat(str(value)).strftime("%d/%m/%Y")
    except Exception:
        return str(value)


def _excel_date(value):
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value)).date()
    except Exception:
        return value


def _sorted(rows: Iterable[dict]) -> list[dict]:
    return sorted(rows, key=lambda x: float(x.get("usage_pct") or 0), reverse=True)


def _thin_border(color="DDE4E8"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _set_col_widths(ws, widths: dict[str, float]):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _history_matrix(rows: list[dict]):
    units = [r.get("unit") for r in rows if r.get("unit")]
    dates = set()
    values = {u: {} for u in units}
    for r in rows:
        unit = r.get("unit")
        for point in r.get("history_series") or []:
            d = str(point.get("date") or "")
            if not d:
                continue
            dates.add(d)
            values.setdefault(unit, {})[d] = float(point.get("usage_pct") or 0) / 100.0
    return sorted(dates), units, values


def _build_excel_summary(wb: Workbook, rows: list[dict], thresholds: dict):
    ws = wb.create_sheet("Resumo Executivo")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    headers = [
        "Sonda", "Periodo Compass", "Terminal", "Plano", "Franquia (GB)", "Consumido (GB)",
        "Disponivel (GB)", "Overage atual (GB)", "Uso (%)", "Status", "Ritmo (GB/dia)",
        "Tendencia", "Pontos historicos", "Dias p/ 100%", "Data estimada 100%", "Fim do ciclo",
        "Projecao fim ciclo (GB)", "Overage projetado (GB)", "Risco projetado", "Metodo projecao",
        "Acao recomendada", "Inicio periodo", "Fim periodo", "Dias periodo", "Atualidade dos dados",
        "Confianca previsao", "Idade dados (dias)"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 42

    for i, r in enumerate(_sorted(rows), 2):
        values = [
            r.get("unit"), r.get("period", ""), r.get("terminal", ""), r.get("plan_name", ""),
            float(r.get("quota_gb") or 0), float(r.get("total_gb") or 0), None, None, None,
            r.get("status", ""), float(r.get("rate_gb_day") or 0), r.get("trend", ""),
            int(r.get("history_points") or 0), r.get("days_to_limit"), _excel_date(r.get("forecast_limit_date")),
            _excel_date(r.get("cycle_end_date")), float(r.get("projected_cycle_end_gb") or 0), None,
            r.get("forecast_risk", ""), r.get("projection_method", ""), r.get("recommended_action", ""),
            _excel_date(r.get("period_start")), _excel_date(r.get("period_end")), int(r.get("period_days") or 0),
            r.get("data_freshness", ""), r.get("forecast_confidence", ""), int(r.get("data_age_days") or 0),
        ]
        for c, v in enumerate(values, 1):
            ws.cell(i, c, v)

        # Simple derived fields stay as formulas so the workbook remains auditable/editable.
        ws.cell(i, 7, f"=MAX(E{i}-F{i},0)")
        ws.cell(i, 8, f"=MAX(F{i}-E{i},0)")
        ws.cell(i, 9, f"=IF(E{i}>0,F{i}/E{i},0)")
        ws.cell(i, 18, f"=MAX(Q{i}-E{i},0)")

        for c in range(1, len(headers) + 1):
            cell = ws.cell(i, c)
            cell.border = Border(bottom=Side(style="thin", color="E8ECEF"))
            cell.alignment = Alignment(vertical="center", wrap_text=c in {2, 3, 4, 12, 19, 20, 21})
        for c in (5, 6, 7, 8, 11, 14, 17, 18):
            ws.cell(i, c).number_format = '#,##0.0'
        ws.cell(i, 9).number_format = "0.0%"
        for c in (15, 16, 22, 23):
            ws.cell(i, c).number_format = "dd/mm/yyyy"

        status = str(r.get("status") or "NORMAL").upper()
        if status in STATUS_COLORS:
            ws.cell(i, 10).fill = PatternFill("solid", fgColor={
                "NORMAL": "E8F5E9", "ATENCAO": "FFF8E1", "CRITICO": "FFF3E0", "EMERGENCIA": "FFEBEE"
            }[status])
            ws.cell(i, 10).font = Font(bold=True, color=STATUS_COLORS[status])
        risk = str(r.get("forecast_risk") or "")
        if risk in RISK_COLORS:
            ws.cell(i, 19).font = Font(bold=True, color=RISK_COLORS[risk])

    ws.auto_filter.ref = ws.dimensions
    widths = [11, 25, 30, 42, 14, 15, 15, 18, 11, 14, 16, 18, 16, 15, 20, 17, 21, 22, 22, 20, 60, 15, 15, 13, 18, 18, 17]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _build_excel_raw(wb: Workbook, rows: list[dict]):
    ws = wb.create_sheet("Dados Coletados")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    headers = [
        "Sonda", "Business Service", "Kit ID(s)", "Kit Name", "Service Line", "Plan Name", "Periodo Compass",
        "Limit (GB)", "Data Boosters (GB)", "Priority (GB)", "Standard (GB)", "% of Limit Compass",
        "Overage calculado (GB)", "Consumido calculado (GB)", "Coletado em", "Inicio periodo", "Fim periodo",
        "Dias periodo", "Arquivo fonte", "SHA256 fonte", "Atualidade", "Confianca previsao"
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, r in enumerate(_sorted(rows), 2):
        values = [
            r.get("unit"), r.get("source_name"), r.get("terminal"), r.get("kit_name"), r.get("service_line"),
            r.get("plan_name"), r.get("period"), float(r.get("quota_gb") or 0), float(r.get("booster_gb") or 0),
            float(r.get("priority_gb") or 0), float(r.get("standard_gb") or 0), float(r.get("portal_usage_pct") or 0) / 100,
            float(r.get("overage_gb") or 0), None, r.get("collected_at"), _excel_date(r.get("period_start")),
            _excel_date(r.get("period_end")), int(r.get("period_days") or 0), r.get("source_file", ""),
            r.get("source_sha256", ""), r.get("data_freshness", ""), r.get("forecast_confidence", "")
        ]
        for c, v in enumerate(values, 1):
            ws.cell(i, c, v)
        ws.cell(i, 14, f"=SUM(I{i}:K{i})")
        for c in range(1, len(headers) + 1):
            ws.cell(i, c).border = Border(bottom=Side(style="thin", color="EDF0F2"))
        for c in (8, 9, 10, 11, 13, 14):
            ws.cell(i, c).number_format = '#,##0.0'
        ws.cell(i, 12).number_format = "0.0%"
        ws.cell(i, 16).number_format = "dd/mm/yyyy"
        ws.cell(i, 17).number_format = "dd/mm/yyyy"
    ws.auto_filter.ref = ws.dimensions
    widths = [11, 32, 31, 31, 25, 48, 25, 14, 18, 15, 15, 18, 20, 22, 21, 15, 15, 12, 36, 34, 16, 18]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _build_excel_history(wb: Workbook, rows: list[dict]):
    ws = wb.create_sheet("Historico")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    headers = ["Data", "Sonda", "Consumido (GB)", "Franquia (GB)", "Uso (%)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(horizontal="center")
    idx = 2
    for r in _sorted(rows):
        for point in r.get("history_series") or []:
            ws.cell(idx, 1, _excel_date(point.get("date")))
            ws.cell(idx, 2, r.get("unit"))
            ws.cell(idx, 3, float(point.get("total_gb") or 0))
            ws.cell(idx, 4, float(point.get("quota_gb") or 0))
            ws.cell(idx, 5, float(point.get("usage_pct") or 0) / 100)
            ws.cell(idx, 1).number_format = "dd/mm/yyyy"
            ws.cell(idx, 5).number_format = "0.0%"
            idx += 1
    ws.auto_filter.ref = ws.dimensions
    _set_col_widths(ws, {"A": 14, "B": 14, "C": 18, "D": 18, "E": 12})


def _build_excel_trend(wb: Workbook, rows: list[dict]):
    ws = wb.create_sheet("Tendencia")
    ws.sheet_view.showGridLines = False
    dates, units, values = _history_matrix(_sorted(rows))
    ws["A1"] = "EVOLUCAO HISTORICA - % DA FRANQUIA"
    ws["A1"].font = Font(size=16, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=DARK)
    end_col = max(len(units) + 1, 8)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)

    ws.cell(3, 1, "Data")
    for c, unit in enumerate(units, 2):
        ws.cell(3, c, unit)
    for c in range(1, len(units) + 2):
        ws.cell(3, c).font = Font(bold=True, color=WHITE)
        ws.cell(3, c).fill = PatternFill("solid", fgColor=TEAL)
        ws.cell(3, c).alignment = Alignment(horizontal="center")

    for r_idx, d in enumerate(dates, 4):
        ws.cell(r_idx, 1, _excel_date(d))
        ws.cell(r_idx, 1).number_format = "dd/mm/yyyy"
        for c_idx, unit in enumerate(units, 2):
            val = values.get(unit, {}).get(d)
            if val is not None:
                ws.cell(r_idx, c_idx, val)
                ws.cell(r_idx, c_idx).number_format = "0.0%"

    chart = LineChart()
    chart.title = "Tendencia de utilizacao da franquia"
    chart.y_axis.title = "% da franquia"
    chart.x_axis.title = "Data"
    chart.height = 10
    chart.width = 22
    if dates and units:
        data = Reference(ws, min_col=2, max_col=1 + len(units), min_row=3, max_row=3 + len(dates))
        cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(dates))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
    ws.add_chart(chart, "A8")

    start = 28
    ws.cell(start, 1, "PREVISAO POR UNIDADE")
    ws.cell(start, 1).font = Font(size=12, bold=True, color=WHITE)
    ws.cell(start, 1).fill = PatternFill("solid", fgColor=DARK)
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=8)
    hdr = ["Sonda", "Ritmo GB/dia", "Tendencia", "Dias p/100%", "Data 100%", "Projecao GB", "Risco", "Metodo"]
    for c, h in enumerate(hdr, 1):
        ws.cell(start + 1, c, h)
        ws.cell(start + 1, c).font = Font(bold=True, color=WHITE)
        ws.cell(start + 1, c).fill = PatternFill("solid", fgColor=TEAL)
    for i, row in enumerate(_sorted(rows), start + 2):
        vals = [
            row.get("unit"), float(row.get("rate_gb_day") or 0), row.get("trend"), row.get("days_to_limit"),
            _excel_date(row.get("forecast_limit_date")), float(row.get("projected_cycle_end_gb") or 0),
            row.get("forecast_risk"), row.get("projection_method"),
        ]
        for c, v in enumerate(vals, 1):
            ws.cell(i, c, v)
        ws.cell(i, 5).number_format = "dd/mm/yyyy"
        for c in (2, 4, 6):
            ws.cell(i, c).number_format = '#,##0.0'
    _set_col_widths(ws, {"A": 14, "B": 18, "C": 20, "D": 16, "E": 16, "F": 18, "G": 22, "H": 20})


def _build_excel_dashboard(wb: Workbook, rows: list[dict], thresholds: dict):
    ws = wb.create_sheet("Painel Executivo", 0)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A9"
    ws.merge_cells("A1:N2")
    ws["A1"] = "STARLINK | PAINEL EXECUTIVO DE CONSUMO E PREVISAO"
    ws["A1"].font = Font(size=20, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=DARK)
    ws["A1"].alignment = Alignment(vertical="center", horizontal="left")
    period = next((r.get("period") for r in rows if r.get("period")), "Periodo informado pelo Compass")
    ws.merge_cells("A3:N3")
    ws["A3"] = f"Periodo Compass: {period} | Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A3"].font = Font(size=10, color="52616B")

    # KPIs use workbook formulas referencing the summary sheet.
    kpis = [
        ("A5", "B5", "CONSUMO TOTAL", "=SUM('Resumo Executivo'!F2:F200)", '0.0 "GB"'),
        ("C5", "D5", "FRANQUIA TOTAL", "=SUM('Resumo Executivo'!E2:E200)", '0.0 "GB"'),
        ("E5", "F5", "OVERAGE ATUAL", "=SUM('Resumo Executivo'!H2:H200)", '0.0 "GB"'),
        ("G5", "H5", "OVERAGE PROJETADO", "=SUM('Resumo Executivo'!R2:R200)", '0.0 "GB"'),
        ("I5", "J5", "ESTOURO PREVISTO", '=COUNTIF(\'Resumo Executivo\'!S2:S200,"ESTOURO PREVISTO")', "0"),
        ("K5", "L5", "ESTOURADO", '=COUNTIF(\'Resumo Executivo\'!S2:S200,"ESTOURADO")', "0"),
        ("M5", "N5", "UNIDADES", "=COUNTA('Resumo Executivo'!A2:A200)", "0"),
    ]
    for start, end, label, formula, numfmt in kpis:
        cell = ws[start]
        row_label = cell.row - 1
        c1, c2 = cell.column, ws[end].column
        ws.merge_cells(start_row=row_label, start_column=c1, end_row=row_label, end_column=c2)
        ws.merge_cells(start_row=cell.row, start_column=c1, end_row=cell.row, end_column=c2)
        lc = ws.cell(row_label, c1)
        lc.value = label
        lc.font = Font(size=8, bold=True, color="52616B")
        lc.fill = PatternFill("solid", fgColor="EAF6F7")
        lc.alignment = Alignment(horizontal="center", vertical="center")
        vc = ws.cell(cell.row, c1)
        vc.value = formula
        vc.number_format = numfmt
        vc.font = Font(size=14, bold=True, color=DARK)
        vc.fill = PatternFill("solid", fgColor="EAF6F7")
        vc.alignment = Alignment(horizontal="center", vertical="center")

    ws["A8"] = "RANKING ATUAL"
    ws["A8"].font = Font(size=12, bold=True, color=WHITE)
    ws["A8"].fill = PatternFill("solid", fgColor=DARK)
    ws.merge_cells("A8:H8")
    headers = ["Sonda", "Uso (%)", "Status", "Ritmo GB/dia", "Tendencia", "Data 100%", "Risco", "Projecao GB"]
    for c, h in enumerate(headers, 1):
        ws.cell(9, c, h)
        ws.cell(9, c).font = Font(bold=True, color=WHITE)
        ws.cell(9, c).fill = PatternFill("solid", fgColor=TEAL)
        ws.cell(9, c).alignment = Alignment(horizontal="center", vertical="center")
    for idx, row in enumerate(_sorted(rows), 10):
        src = idx - 8
        refs = [
            f"='Resumo Executivo'!A{src}", f"='Resumo Executivo'!I{src}", f"='Resumo Executivo'!J{src}",
            f"='Resumo Executivo'!K{src}", f"='Resumo Executivo'!L{src}", f"='Resumo Executivo'!O{src}",
            f"='Resumo Executivo'!S{src}", f"='Resumo Executivo'!Q{src}",
        ]
        for c, ref in enumerate(refs, 1):
            ws.cell(idx, c, ref)
            ws.cell(idx, c).border = Border(bottom=Side(style="thin", color="E3E8EB"))
            ws.cell(idx, c).alignment = Alignment(horizontal="center" if c != 1 else "left")
        ws.cell(idx, 2).number_format = "0.0%"
        ws.cell(idx, 6).number_format = "dd/mm/yyyy"
        ws.cell(idx, 4).number_format = '#,##0.0'
        ws.cell(idx, 8).number_format = '#,##0.0'

    if rows:
        end_row = 9 + len(rows)
        ws.conditional_formatting.add(f"B10:B{end_row}", CellIsRule(operator="greaterThanOrEqual", formula=[str(thresholds["emergency"] / 100)], fill=PatternFill("solid", fgColor="FDECEC")))
        ws.conditional_formatting.add(f"B10:B{end_row}", CellIsRule(operator="between", formula=[str(thresholds["critical"] / 100), str((thresholds["emergency"] - 0.01) / 100)], fill=PatternFill("solid", fgColor="FFF1E5")))

    chart = BarChart()
    chart.type = "bar"
    chart.title = "Utilizacao atual da franquia"
    chart.y_axis.title = "Unidade"
    chart.x_axis.title = "% da franquia"
    chart.height = 7.2
    chart.width = 12.5
    if rows:
        data = Reference(ws, min_col=2, min_row=9, max_row=9 + len(rows))
        cats = Reference(ws, min_col=1, min_row=10, max_row=9 + len(rows))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
    ws.add_chart(chart, "J8")

    chart2 = BarChart()
    chart2.type = "col"
    chart2.title = "Projecao fim do ciclo x franquia"
    chart2.y_axis.title = "GB"
    chart2.x_axis.title = "Unidade"
    chart2.height = 7.2
    chart2.width = 12.5
    if rows:
        rs = wb["Resumo Executivo"]
        data = Reference(rs, min_col=5, max_col=5, min_row=1, max_row=1 + len(rows))
        chart2.add_data(data, titles_from_data=True)
        data2 = Reference(rs, min_col=17, max_col=17, min_row=1, max_row=1 + len(rows))
        chart2.add_data(data2, titles_from_data=True)
        cats = Reference(rs, min_col=1, min_row=2, max_row=1 + len(rows))
        chart2.set_categories(cats)
    ws.add_chart(chart2, "J23")
    _set_col_widths(ws, {"A": 13, "B": 12, "C": 15, "D": 16, "E": 18, "F": 16, "G": 22, "H": 17, "I": 3, "J": 12, "K": 12, "L": 12, "M": 12, "N": 12})


def _build_excel_parameters(wb: Workbook, config: dict):
    ws = wb.create_sheet("Parametros")
    ws.sheet_state = "hidden"
    thresholds = config.get("thresholds", {})
    history = config.get("history", {})
    projection = config.get("projection", {})
    rows = [
        ("Parametro", "Valor"),
        ("warning", float(thresholds.get("warning", 70)) / 100),
        ("critical", float(thresholds.get("critical", 85)) / 100),
        ("emergency", float(thresholds.get("emergency", 95)) / 100),
        ("cycle_start_day", int(projection.get("cycle_start_day", 1))),
        ("history_lookback_days", int(history.get("lookback_days", 90))),
        ("min_points_for_history_rate", int(history.get("min_points_for_history_rate", 2))),
    ]
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in range(2, 5):
        ws.cell(r, 2).number_format = "0.0%"


def generate_excel(rows, config=None):
    config = config or {}
    thresholds = config.get("thresholds", {"warning": 70, "critical": 85, "emergency": 95})
    date_tag = datetime.now().strftime("%Y-%m-%d")
    path = OUTPUT / f"Relatorio_Consumo_Starlink_{date_tag}.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    _build_excel_summary(wb, rows, thresholds)
    _build_excel_raw(wb, rows)
    _build_excel_history(wb, rows)
    _build_excel_trend(wb, rows)
    _build_excel_parameters(wb, config)
    _build_excel_dashboard(wb, rows, thresholds)
    wb.active = 0
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass
    wb.save(path)
    return path


def _pdf_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="ExecTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=19, leading=22, textColor=colors.HexColor(f"#{DARK}"), alignment=TA_LEFT, spaceAfter=2))
    styles.add(ParagraphStyle(name="ExecSub", parent=styles["Normal"], fontName="Helvetica", fontSize=8.5, leading=11, textColor=colors.HexColor("#60727C"), alignment=TA_LEFT))
    styles.add(ParagraphStyle(name="Section", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=12, leading=14, textColor=colors.HexColor(f"#{DARK}"), spaceBefore=4, spaceAfter=6))
    styles.add(ParagraphStyle(name="Small", parent=styles["Normal"], fontSize=7.3, leading=9.0, textColor=colors.HexColor(f"#{TEXT}")))
    styles.add(ParagraphStyle(name="Tiny", parent=styles["Normal"], fontSize=6.3, leading=7.5, textColor=colors.HexColor(f"#{TEXT}")))
    styles.add(ParagraphStyle(name="KPI", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=13, leading=15, alignment=TA_CENTER, textColor=colors.HexColor(f"#{DARK}")))
    styles.add(ParagraphStyle(name="KPILabel", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=6.6, leading=8, alignment=TA_CENTER, textColor=colors.HexColor("#60727C")))
    return styles


def _pdf_header(story, styles, company: str, period: str):
    logo = ASSETS / "logo.png"
    if logo.exists():
        try:
            logo_flow = Image(str(logo), width=30 * mm, height=12 * mm, kind="proportional")
        except Exception:
            logo_flow = Paragraph(f"<b>{company}</b>", styles["ExecTitle"])
    else:
        logo_flow = Paragraph(f"<b>{company}</b>", styles["ExecTitle"])
    title = [
        Paragraph("Relatorio Executivo - Consumo Starlink", styles["ExecTitle"]),
        Paragraph(f"Periodo Compass: {period} | Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["ExecSub"]),
    ]
    header = Table([[logo_flow, title]], colWidths=[42 * mm, 210 * mm])
    header.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("LINEBELOW", (0, 0), (-1, -1), 1.2, colors.HexColor(f"#{TEAL}")), ("BOTTOMPADDING", (0, 0), (-1, -1), 7)]))
    story.append(header)
    story.append(Spacer(1, 7))


def _kpi_card(value: str, label: str, styles, alert=False):
    bg = colors.HexColor("#FDECEC") if alert else colors.HexColor("#EAF6F7")
    return Table([[Paragraph(value, styles["KPI"])], [Paragraph(label, styles["KPILabel"])]], colWidths=[39 * mm], rowHeights=[9 * mm, 7 * mm], style=TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), bg), ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#C5D8DC")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))


def _usage_chart(rows: list[dict]) -> Drawing:
    width = 250 * mm
    height = 45 * mm
    drawing = Drawing(width, height)
    chart = HorizontalBarChart()
    chart.x = 34 * mm
    chart.y = 5 * mm
    chart.width = 205 * mm
    chart.height = height - 10 * mm
    values = [min(float(r.get("usage_pct") or 0), 130.0) for r in rows]
    chart.data = [values]
    chart.categoryAxis.categoryNames = [str(r.get("unit") or "-") for r in rows]
    chart.categoryAxis.labels.fontSize = 7
    chart.valueAxis.valueMin = 0
    chart.valueAxis.valueMax = max(110, int(max(values or [100]) / 10 + 1) * 10)
    chart.valueAxis.valueStep = 20
    chart.valueAxis.labels.fontSize = 7
    chart.valueAxis.labelTextFormat = "%d%%"
    chart.bars[0].fillColor = colors.HexColor(f"#{TEAL}")
    chart.bars[0].strokeColor = None
    drawing.add(chart)
    return drawing


def _history_chart(rows: list[dict]) -> Drawing:
    dates, units, values = _history_matrix(rows)
    width, height = 255 * mm, 68 * mm
    d = Drawing(width, height)
    left, right, bottom, top = 18 * mm, 6 * mm, 15 * mm, 8 * mm
    plot_w, plot_h = width - left - right, height - bottom - top
    if not dates or not units:
        d.add(String(width / 2, height / 2, "Historico insuficiente para o grafico de tendencia.", textAnchor="middle", fontSize=9, fillColor=colors.HexColor("#60727C")))
        return d

    max_pct = max([100.0] + [float(v) * 100 for unit in units for v in values.get(unit, {}).values()])
    y_max = max(110.0, min(150.0, ((int(max_pct) // 10) + 1) * 10.0))
    for pct in range(0, int(y_max) + 1, 25):
        y = bottom + plot_h * pct / y_max
        d.add(Line(left, y, left + plot_w, y, strokeColor=colors.HexColor("#E1E7EA"), strokeWidth=0.4))
        d.add(String(left - 2 * mm, y - 2, f"{pct}%", textAnchor="end", fontSize=6.5, fillColor=colors.HexColor("#60727C")))
    d.add(Line(left, bottom, left, bottom + plot_h, strokeColor=colors.HexColor("#90A4AE"), strokeWidth=0.7))
    d.add(Line(left, bottom, left + plot_w, bottom, strokeColor=colors.HexColor("#90A4AE"), strokeWidth=0.7))

    def x_for(idx):
        return left + (plot_w * idx / max(len(dates) - 1, 1))

    def y_for(pct):
        return bottom + plot_h * min(max(pct, 0), y_max) / y_max

    for ui, unit in enumerate(units):
        col = colors.HexColor(f"#{CHART_COLORS[ui % len(CHART_COLORS)]}")
        pts = []
        for idx, day in enumerate(dates):
            val = values.get(unit, {}).get(day)
            if val is not None:
                pts.append((x_for(idx), y_for(float(val) * 100)))
        for a, b in zip(pts, pts[1:]):
            d.add(Line(a[0], a[1], b[0], b[1], strokeColor=col, strokeWidth=1.4))
        for x, y in pts:
            d.add(Circle(x, y, 1.3, fillColor=col, strokeColor=None))

    label_indices = sorted(set([0, len(dates) // 2, len(dates) - 1]))
    for idx in label_indices:
        try:
            label = datetime.fromisoformat(dates[idx]).strftime("%d/%m")
        except Exception:
            label = dates[idx]
        d.add(String(x_for(idx), bottom - 4 * mm, label, textAnchor="middle", fontSize=6.5, fillColor=colors.HexColor("#60727C")))

    legend_x, legend_y = left, height - 4 * mm
    for ui, unit in enumerate(units):
        col = colors.HexColor(f"#{CHART_COLORS[ui % len(CHART_COLORS)]}")
        x = legend_x + (ui % 5) * 48 * mm
        y = legend_y - (ui // 5) * 5 * mm
        d.add(Rect(x, y - 2, 4, 4, fillColor=col, strokeColor=None))
        d.add(String(x + 6, y - 2, str(unit), fontSize=6.5, fillColor=colors.HexColor(f"#{TEXT}")))
    return d


def _footer(canvas, doc):
    canvas.saveState()
    canvas.setStrokeColor(colors.HexColor("#D8E0E4"))
    canvas.line(18 * mm, 10 * mm, 279 * mm, 10 * mm)
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(colors.HexColor("#71808A"))
    canvas.drawString(18 * mm, 6 * mm, f"Starlink Consumption Agent v{APP_VERSION} | periodo Compass + historico SQLite + previsao")
    canvas.drawRightString(279 * mm, 6 * mm, f"Pagina {doc.page}")
    canvas.restoreState()


def generate_pdf(rows, config=None, output_path=None):
    config = config or {}
    company = config.get("company", "Foresea")
    date_tag = datetime.now().strftime("%Y-%m-%d")
    path = Path(output_path) if output_path else OUTPUT / f"Relatorio_Executivo_Starlink_{date_tag}.pdf"
    path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(str(path), pagesize=landscape(A4), rightMargin=16 * mm, leftMargin=16 * mm, topMargin=12 * mm, bottomMargin=15 * mm)
    styles = _pdf_styles()
    sorted_rows = _sorted(rows)
    period = next((r.get("period") for r in rows if r.get("period")), "Periodo informado pelo Compass")
    total_usage = sum(float(r.get("total_gb") or 0) for r in rows)
    total_quota = sum(float(r.get("quota_gb") or 0) for r in rows)
    total_overage = sum(float(r.get("overage_gb") or 0) for r in rows)
    projected_overage = sum(float(r.get("projected_overage_gb") or 0) for r in rows)
    forecast_breach = sum(1 for r in rows if str(r.get("forecast_risk")) in {"ESTOURADO", "ESTOURO PREVISTO"})
    urgent = sum(1 for r in rows if str(r.get("status")) in {"CRITICO", "EMERGENCIA"})

    story = []
    _pdf_header(story, styles, company, period)
    kpis = Table([[
        _kpi_card(_fmt_gb(total_usage), "CONSUMO TOTAL", styles),
        _kpi_card(_fmt_gb(total_quota), "FRANQUIA TOTAL", styles),
        _kpi_card(_fmt_gb(total_overage), "OVERAGE ATUAL", styles, alert=total_overage > 0),
        _kpi_card(_fmt_gb(projected_overage), "OVERAGE PROJETADO", styles, alert=projected_overage > 0),
        _kpi_card(str(forecast_breach), "ESTOURO ATUAL/PREVISTO", styles, alert=forecast_breach > 0),
        _kpi_card(str(urgent), "CRITICAS/EMERGENCIA", styles, alert=urgent > 0),
    ]], colWidths=[42 * mm] * 6)
    kpis.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
    story.append(kpis)
    story.append(Spacer(1, 7))
    story.append(Paragraph("Ranking de utilizacao atual", styles["Section"]))
    if sorted_rows:
        story.append(_usage_chart(sorted_rows))
    story.append(Spacer(1, 3))

    data = [["Sonda", "Franquia", "Consumido", "Uso", "Status", "Ritmo", "Tendencia", "Risco projetado", "Data 100%"]]
    for r in sorted_rows:
        data.append([
            r.get("unit", "-"), _fmt_gb(r.get("quota_gb")), _fmt_gb(r.get("total_gb")), f"{float(r.get('usage_pct') or 0):.1f}%",
            r.get("status", "-"), f"{float(r.get('rate_gb_day') or 0):.1f} GB/d", r.get("trend", "-"),
            r.get("forecast_risk", "-"), _fmt_date(r.get("forecast_limit_date")),
        ])
    table = Table(data, repeatRows=1, colWidths=[17*mm, 24*mm, 25*mm, 17*mm, 23*mm, 24*mm, 31*mm, 39*mm, 29*mm])
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{DARK}")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 7.1),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("ALIGN", (1, 1), (-1, -1), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LINEBELOW", (0, 1), (-1, -1), 0.25, colors.HexColor("#DCE3E7")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F9FA")]),
    ]
    for idx, r in enumerate(sorted_rows, 1):
        status = str(r.get("status") or "NORMAL")
        fill = STATUS_COLORS.get(status, GREEN)
        style_cmds += [("BACKGROUND", (4, idx), (4, idx), colors.HexColor(f"#{fill}")), ("TEXTCOLOR", (4, idx), (4, idx), colors.white), ("FONTNAME", (4, idx), (4, idx), "Helvetica-Bold")]
        risk = str(r.get("forecast_risk") or "CONTROLADO")
        rfill = RISK_COLORS.get(risk, GREEN)
        style_cmds += [("TEXTCOLOR", (7, idx), (7, idx), colors.HexColor(f"#{rfill}")), ("FONTNAME", (7, idx), (7, idx), "Helvetica-Bold")]
    table.setStyle(TableStyle(style_cmds))
    story.append(table)

    # Page 2 - trend and forecast.
    story.append(PageBreak())
    _pdf_header(story, styles, company, period)
    story.append(Paragraph("Tendencia historica - percentual da franquia", styles["Section"]))
    story.append(_history_chart(sorted_rows))
    story.append(Spacer(1, 5))
    story.append(Paragraph("Previsao por unidade", styles["Section"]))
    forecast = [["Sonda", "Ritmo GB/dia", "Tendencia", "Pontos", "Ciclo termina", "Dias p/100%", "Data 100%", "Projecao fim", "Overage proj.", "Risco", "Metodo"]]
    for r in sorted_rows:
        forecast.append([
            r.get("unit", "-"), f"{float(r.get('rate_gb_day') or 0):.1f}", r.get("trend", "-"), str(r.get("history_points") or 0),
            _fmt_date(r.get("cycle_end_date")), "-" if r.get("days_to_limit") is None else f"{float(r.get('days_to_limit')):.1f}",
            _fmt_date(r.get("forecast_limit_date")), _fmt_gb(r.get("projected_cycle_end_gb")), _fmt_gb(r.get("projected_overage_gb")),
            r.get("forecast_risk", "-"), r.get("projection_method", "-"),
        ])
    ft = Table(forecast, repeatRows=1, colWidths=[15*mm, 22*mm, 28*mm, 15*mm, 25*mm, 22*mm, 25*mm, 27*mm, 26*mm, 35*mm, 28*mm])
    ft.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{TEAL}")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 6.5),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("ALIGN", (1, 1), (-1, -1), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3), ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LINEBELOW", (0, 1), (-1, -1), 0.25, colors.HexColor("#DCE3E7")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F9FA")]),
    ]))
    story.append(ft)
    story.append(Spacer(1, 6))
    story.append(Paragraph("Metodo: com pelo menos 2 snapshots validos no ciclo, o agente calcula o ritmo pela tendencia historica. Sem historico suficiente, usa a media do periodo exportado pelo Compass quando disponivel; caso contrario, usa a media do ciclo. Quedas bruscas do acumulado sao tratadas como reinicio de serie para evitar previsoes incorretas.", styles["Small"]))

    # Page 3 - technical/audit details.
    story.append(PageBreak())
    _pdf_header(story, styles, company, period)
    story.append(Paragraph("Detalhamento tecnico e recomendacoes", styles["Section"]))
    detail = [["Sonda", "Terminal", "Service Line", "Plano", "Priority", "Booster", "Standard", "Uso Compass", "Acao recomendada"]]
    for r in sorted_rows:
        detail.append([
            r.get("unit", "-"), Paragraph(str(r.get("terminal") or "-"), styles["Tiny"]),
            Paragraph(str(r.get("service_line") or "-"), styles["Tiny"]), Paragraph(str(r.get("plan_name") or "-"), styles["Tiny"]),
            _fmt_gb(r.get("priority_gb")), _fmt_gb(r.get("booster_gb")), _fmt_gb(r.get("standard_gb")),
            f"{float(r.get('portal_usage_pct') or 0):.2f}%", Paragraph(str(r.get("recommended_action") or "-"), styles["Tiny"]),
        ])
    dt = Table(detail, repeatRows=1, colWidths=[14*mm, 36*mm, 33*mm, 70*mm, 20*mm, 20*mm, 20*mm, 23*mm, 44*mm])
    dt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{DARK}")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 6.4),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("ALIGN", (4, 1), (7, -1), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3), ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LINEBELOW", (0, 1), (-1, -1), 0.25, colors.HexColor("#DCE3E7")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F9FA")]),
    ]))
    story.append(dt)
    story.append(Spacer(1, 8))
    story.append(Paragraph("Observacao: a previsao depende do dia de inicio do ciclo configurado em projection.cycle_start_day. Ajuste esse parametro ao ciclo real de faturamento Starlink/Compass para obter previsoes de fim de ciclo mais precisas.", styles["Small"]))

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return path
